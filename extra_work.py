import os
import re
import math
from datetime import date
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

extra_work_types = ['Обычная', 'Повышенная']
type_break_time = '22:00'
holidays = [
    "01.01.2023"
    "01.02.2023"
    "01.03.2023"
    "01.04.2023"
    "01.05.2023"
    "01.06.2023"
    "01.07.2023"
    "01.08.2023"
    "02.22.2023"
    "02.23.2023"
    "02.24.2023"
    "03.07.2023"
    "03.08.2023"
    "05.01.2023"
    "05.08.2023"
    "05.09.2023"
    "06.12.2023"
    "11.03.2023"
    "11.04.2023"
    "11.06.2023"
]


def extract_table(doc_path: str, table_index: int=0):
    # Load the Word document
    doc = Document(doc_path)

    # Get the specified table from the document
    tables = doc.tables
    if table_index < len(tables):
        table = tables[table_index]

        # Extract data from the table
        data = []
        for row in table.rows:
            row_data = [cell.text for cell in row.cells]
            data.append(row_data)

        return data


def create_excel(doc_path: str, table_data: list):
    wb = Workbook()
    worksheet = wb.active

    columns = ['Сотрудник', 'Способ компенсации', 'Период работы', 'Задача JIRA']
    _type_break_time = convert_time_to_minutes(type_break_time)

    # Добавляем столбцы дат
    i = 1
    while i < len(table_data):
        row = table_data[i]
        i += 1
        date_str = format_date_str(row[2].strip())

        if (date_str not in columns):
            columns.insert(1, date_str)

    # Записываем первый ряд с заголовками
    i = 0
    while i < len(columns):
        i += 1
        worksheet.cell(1, i).value = columns[i - 1]
    
    # Записываем в ячейки
    i = 1
    row_index = i
    while i < len(table_data):
        i += 1
        row_index += 1
        row = table_data[i - 1]
        date_str = format_date_str(row[2].strip())
        dateIndex = columns.index(date_str)
        day, month, year = date_str.split('.')
        date_obj = date(int(year), int(month), int(day))

        start_time, end_time = re.sub("[\s]", '', row[4]).split('-')
        _start_time = convert_time_to_minutes(start_time)
        _end_time = convert_time_to_minutes(end_time)
        
        worksheet.cell(row_index, 1).value = row[0] # ФИО
        worksheet.cell(row_index, dateIndex + 1).value = row[3] # Количество часов
        worksheet.cell(row_index, len(columns) - 2).value = extra_work_types[0] # Тип компенсации
        worksheet.cell(row_index, len(columns) - 1).value = f'{start_time} - {end_time}' # Период работы
        worksheet.cell(row_index, len(columns)).value = row[5] # Задача JIRA

        if _start_time >= _type_break_time or date_obj.isoweekday() in [6, 7] or date_str in holidays:
            worksheet.cell(row_index, len(columns) - 2).value = extra_work_types[1]
        elif _end_time > _type_break_time:
            worksheet.cell(row_index, dateIndex + 1).value = convert_minutes_to_time(_type_break_time - _start_time) # Количество часов
            worksheet.cell(row_index, len(columns) - 1).value = f'{start_time} - {type_break_time}' # Период работы
            row_index += 1
            worksheet.cell(row_index, 1).value = row[0] # ФИО
            worksheet.cell(row_index, dateIndex + 1).value = convert_minutes_to_time(_end_time - _type_break_time) # Количество часов
            worksheet.cell(row_index, len(columns) - 2).value = extra_work_types[1] # Тип компенсации
            worksheet.cell(row_index, len(columns) - 1).value = f'{type_break_time} - {end_time}' # Период работы
            worksheet.cell(row_index, len(columns)).value = row[5] # Задача JIRA
    
    adjust_worksheet_width(worksheet)
    worksheet_solors(worksheet)
    wb.save(doc_path)


def adjust_worksheet_width(worksheet):
    column_widths = []
    for row in worksheet.rows:
        for i, cell in enumerate(row):
            if type(cell.value) != str:
                continue
            width = len(cell.value)
            
            if len(column_widths) - 1 < i:
                column_widths.append(width)
            elif column_widths[i] < width:
                column_widths[i] = width

    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
        worksheet.column_dimensions[get_column_letter(i)].width = column_width + 1


def worksheet_solors(worksheet):
    for rows in worksheet.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type = "solid")


def convert_time_to_minutes(time: str) -> int:
    hours, minutes = time.split(':')
    return int(hours) * 60 + int(minutes)


def convert_minutes_to_time(minutes: int) -> str:
    hours = math.floor(minutes / 60)
    minutes = minutes % 60
    time = str(hours)

    if minutes > 0:
        time += f':{minutes}'
    
    return time

def format_date_str(date_str: str) -> str:
    day, month, year = date_str.split('.')
    if len(year) == 2:
        year = '20' + year
    
    return f'{day}.{month}.{year}'

def docx_to_xlsx(input_path: str, export_path: str):
    table_data = extract_table(input_path, 1)

    if table_data:
        create_excel(export_path, table_data)
        print('Сохранено в файл ' + export_path)
    else:
        print('Таблица не найдена')

if __name__ == "__main__":
    for filename in os.listdir('src'):
        basename = filename.split('.')
        basename.pop()
        basename = '.'.join(basename)
        docx_to_xlsx(f'src/{filename}', f'dist/{basename}.xlsx')
